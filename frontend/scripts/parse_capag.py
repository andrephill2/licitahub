import openpyxl, json, unicodedata, sys

def norm(s):
    s = str(s or "").strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower()

XLSX = r"C:\Users\SOSDOC~1\AppData\Local\Temp\claude\C--Users-SOS-DOCS\09ebec46-6d6b-4fb8-9ba6-856ec5322b6d\scratchpad\capag_municipios.xlsx"

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active

if "--inspect" in sys.argv:
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(i, [str(c)[:40] if c is not None else "None" for c in row[:15]])
        if i >= 6: break
    wb.close()
    sys.exit(0)

# Estrutura confirmada pela inspeção:
# Col 0: Código Município Completo (IBGE 7 dígitos)
# Col 1: Nome_Município
# Col 2: UF
# Col 3: CAPAG (A+/A/B+/B/C/D/SC)
# Cabecalho na linha 2 (índice 2), dados a partir da linha 3

capag_data = []

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 3:
        continue  # pula as 2 linhas de totais + linha de cabeçalho

    ibge = str(row[0]).strip() if row[0] is not None else ""
    mun  = str(row[1]).strip() if row[1] is not None else ""
    uf   = str(row[2]).strip().upper() if row[2] is not None else ""
    capag = str(row[3]).strip().upper() if row[3] is not None else ""

    if not mun or not uf or not capag or capag in ("CAPAG", "NONE", ""):
        continue

    # Normaliza nota: A+/A→A, B+/B→B, C→C, D→D, SC→SC
    nota_base = capag.rstrip("+")

    entry = {
        "uf": uf,
        "mun": mun,
        "mun_norm": norm(mun),
        "capag": capag,      # nota completa (ex: A+, B+)
        "nota": nota_base,   # nota base (A/B/C/D/SC)
    }
    if ibge and ibge not in ("None", "nan", ""):
        entry["ibge"] = ibge[:7]
    capag_data.append(entry)

wb.close()
print(f"Total registros: {len(capag_data)}", file=sys.stderr)
print(f"Exemplo: {capag_data[:3]}", file=sys.stderr)

out = r"C:\Users\SOS DOCS\licitahub\frontend\src\data\capag.json"
with open(out, "w", encoding="utf-8") as f:
    json.dump(capag_data, f, ensure_ascii=False, separators=(",", ":"))
print(f"Salvo em {out}", file=sys.stderr)
print("OK")
